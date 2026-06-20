from __future__ import annotations

import argparse
import asyncio
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from luoying_bot.capabilities.knowledge_base.artifacts import MarkdownArtifactStore
from luoying_bot.capabilities.knowledge_base.embeddings import OpenAICompatibleEmbeddingProvider
from luoying_bot.capabilities.knowledge_base.entities import (
    GLOBAL_ENTITY_SPACE_ID,
    normalize_entity_text,
    stable_entity_id,
    stable_search_item_id,
)
from luoying_bot.capabilities.knowledge_base.postgres_store import IndexedDocument, PostgresKnowledgeStore
from luoying_bot.capabilities.knowledge_base.quality import MarkdownQualityChecker
from luoying_bot.config import settings


BASE_URL = "https://zsdata.whu.edu.cn/wzgl/wxmini"
SOURCE_URL = "https://zsdata.whu.edu.cn/public/wzgl/#/jhcx"
SITE_URL = "https://zsdata.whu.edu.cn/public/wzgl/"
SITE_ID = "whu_admissions"
SPACE_ID = "whu"
STRONG_FOUNDATION_XLSX = Path("docs/2025分省（区）录取分数及位次 - 挂网 - 最新.xlsx")
STRONG_FOUNDATION_PROGRAM = "数学与应用数学（智能科学）强基计划"
ENTITY_SEED_PATH = Path("knowledge/seeds/admissions_entities.json")


async def main() -> None:
    parser = argparse.ArgumentParser(description="Import real WHU admission plan and score data into the KB")
    parser.add_argument("--province", default="", help="Optional province filter, for example 湖北")
    parser.add_argument("--year", default="", help="Optional year filter, for example 2025")
    parser.add_argument("--subject-type", default="", help="Optional subject filter, for example 物理类")
    parser.add_argument(
        "--strong-foundation-xlsx",
        default=str(STRONG_FOUNDATION_XLSX),
        help="Excel source for strong foundation score data",
    )
    parser.add_argument(
        "--entity-seed",
        default=str(ENTITY_SEED_PATH),
        help="Curated entity and alias seed file",
    )
    args = parser.parse_args()

    store = PostgresKnowledgeStore(
        settings.kb_database_url,
        embedding_provider=OpenAICompatibleEmbeddingProvider(
            base_url=settings.kb_embedding_base_url,
            api_key=settings.kb_embedding_api_key,
            model=settings.kb_embedding_model,
            batch_size=settings.kb_embedding_batch_size,
        ),
        embedding_dimensions=settings.kb_embedding_dimensions,
    )
    await store.ensure_schema()

    async with httpx.AsyncClient(verify=False, trust_env=False, timeout=30) as client:
        plan_rows, plan_raw = await fetch_rows(
            client,
            data_type="zsjh",
            province=args.province,
            year=args.year,
            subject_type=args.subject_type,
        )
        score_rows, score_raw = await fetch_rows(
            client,
            data_type="lnfs",
            province=args.province,
            year=args.year,
            subject_type=args.subject_type,
        )
        site_payload = await fetch_admission_site_payload(client)

    normalized_plans = normalize_plan_rows(plan_rows)
    normalized_scores = normalize_score_rows(score_rows)
    normalized_strong_foundation_scores = normalize_strong_foundation_rows(
        Path(args.strong_foundation_xlsx),
        year_filter=args.year,
        province_filter=args.province,
        subject_type_filter=args.subject_type,
    )
    imported_plans = await store.upsert_admission_plans(normalized_plans)
    imported_scores = await store.upsert_admission_scores(normalized_scores)
    imported_strong_foundation_scores = await store.upsert_admission_strong_foundation_scores(
        normalized_strong_foundation_scores
    )
    content_categories = normalize_content_categories(site_payload["content_categories"])
    content_articles = normalize_content_articles(site_payload["content_articles"])
    academic_payload = normalize_academic_structure(site_payload["academic_units"])
    media_categories = normalize_media_categories(site_payload["media_categories"])
    media_items = normalize_media_items(site_payload["media_items"])
    imported_content_categories = await store.upsert_admission_content_categories(content_categories)
    imported_articles = await store.upsert_admission_articles(content_articles)
    imported_academic_units = await store.upsert_academic_units(academic_payload["units"])
    imported_schools = await store.upsert_admission_schools(academic_payload["schools"])
    imported_majors = await store.upsert_majors(academic_payload["majors"])
    imported_media_categories = await store.upsert_admission_content_categories(media_categories)
    imported_media_items = await store.upsert_admission_media_items(media_items)
    entity_payload = build_admission_entities(
        plans=normalized_plans,
        scores=normalized_scores,
        strong_foundation_scores=normalized_strong_foundation_scores,
        site_majors=academic_payload["majors"],
        seed_path=Path(args.entity_seed),
    )
    await store.clear_kb_entities(SPACE_ID)
    imported_entities = await store.upsert_kb_entities(entity_payload["entities"])
    imported_entity_aliases = await store.upsert_kb_entity_aliases(entity_payload["aliases"])
    imported_entity_relations = await store.upsert_kb_entity_relations(entity_payload["relations"])
    search_items = build_search_items(
        entity_payload=entity_payload,
        plans=normalized_plans,
        scores=normalized_scores,
        strong_foundation_scores=normalized_strong_foundation_scores,
    )
    await store.clear_kb_search_items(SPACE_ID)
    imported_search_items = await store.upsert_kb_search_items(search_items)

    artifact_store = MarkdownArtifactStore(settings.kb_artifact_root)
    artifact_store.write_source(
        {
            "site_id": SITE_ID,
            "name": "武汉大学本科招生数据",
            "base_url": SOURCE_URL,
            "space_id": SPACE_ID,
            "allowed_domains": ["zsdata.whu.edu.cn"],
            "entry_urls": [SOURCE_URL],
            "updated_at": now_iso(),
            "data_api": f"{BASE_URL}/api/front/lqxx/getList",
            "content_api": f"{BASE_URL}/api/front/commcfg",
        }
    )
    plan_artifact = write_dataset_artifact(
        artifact_store,
        title="武汉大学本科招生计划",
        url=f"{SOURCE_URL}?dataset=zsjh",
        rows=normalized_plans,
        raw_payload=plan_raw,
        row_markdown=plan_row_markdown,
    )
    score_artifact = write_dataset_artifact(
        artifact_store,
        title="武汉大学历年录取分数",
        url=f"{SOURCE_URL}?dataset=lnfs",
        rows=normalized_scores,
        raw_payload=score_raw,
        row_markdown=score_row_markdown,
    )
    strong_foundation_artifact = write_dataset_artifact(
        artifact_store,
        title="武汉大学强基计划录取分数",
        url=f"{SOURCE_URL}?dataset=strong_foundation",
        rows=normalized_strong_foundation_scores,
        raw_payload=[
            {
                "source_file": str(Path(args.strong_foundation_xlsx)),
                "program_name": STRONG_FOUNDATION_PROGRAM,
                "rows": normalized_strong_foundation_scores,
            }
        ],
        row_markdown=strong_foundation_row_markdown,
    )
    await index_artifact(store, plan_artifact)
    await index_artifact(store, score_artifact)
    if normalized_strong_foundation_scores:
        await index_artifact(store, strong_foundation_artifact)
    article_artifacts = write_article_artifacts(artifact_store, content_articles)
    media_artifact = write_media_catalog_artifact(artifact_store, media_items)
    academic_artifact = write_academic_structure_artifact(artifact_store, academic_payload)
    for artifact in [*article_artifacts, media_artifact, academic_artifact]:
        await index_artifact(store, artifact)

    print(
        json.dumps(
            {
                "ok": True,
                "plans_imported": imported_plans,
                "scores_imported": imported_scores,
                "strong_foundation_scores_imported": imported_strong_foundation_scores,
                "content_categories_imported": imported_content_categories,
                "articles_imported": imported_articles,
                "academic_units_imported": imported_academic_units,
                "schools_imported": imported_schools,
                "majors_imported": imported_majors,
                "media_categories_imported": imported_media_categories,
                "media_items_imported": imported_media_items,
                "entities_imported": imported_entities,
                "entity_aliases_imported": imported_entity_aliases,
                "entity_relations_imported": imported_entity_relations,
                "search_items_imported": imported_search_items,
                "plan_raw_requests": len(plan_raw),
                "score_raw_requests": len(score_raw),
                "article_artifacts_indexed": len(article_artifacts),
                "artifact_root": str(settings.kb_artifact_root),
                "database_url": settings.kb_database_url,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


async def fetch_rows(
    client: httpx.AsyncClient,
    *,
    data_type: str,
    province: str,
    year: str,
    subject_type: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    type_payload = await post_api(client, "/api/front/lqxx/getType", {"type": data_type})
    type_map = type_payload.get("typeMap")
    if not isinstance(type_map, dict):
        raise ValueError(f"{data_type} getType returned no typeMap")
    rows: list[dict[str, Any]] = []
    raw_payloads: list[dict[str, Any]] = [{"endpoint": "getType", "data_type": data_type, "payload": type_payload}]
    seen_queries: set[tuple[str, str, str, str]] = set()
    for query in iter_queries(type_map, data_type=data_type, province=province, year=year, subject_type=subject_type):
        query_key = (query["sf"], query["nf"], query["klmc"], query["xqmc"])
        if query_key in seen_queries:
            continue
        seen_queries.add(query_key)
        payload = await post_api(client, "/api/front/lqxx/getList", query)
        raw_payloads.append({"endpoint": "getList", "data_type": data_type, "query": query, "payload": payload})
        payload_rows = payload.get("list") or payload.get("data") or []
        if not isinstance(payload_rows, list):
            continue
        rows.extend(row for row in payload_rows if isinstance(row, dict))
    return rows, raw_payloads


async def post_api(client: httpx.AsyncClient, path: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = await client.post(
        f"{BASE_URL}{path}",
        json=payload,
        headers={
            "Content-Type": "application/json;charset=utf-8",
        },
    )
    response.raise_for_status()
    data = response.json()
    if int(data.get("code") or 0) != 200:
        raise ValueError(f"{path} failed: {data}")
    return data


async def post_form_api(client: httpx.AsyncClient, path: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = await client.post(
        f"{BASE_URL}{path}",
        data={key: value for key, value in payload.items() if value is not None},
        headers={"Content-Type": "application/x-www-form-urlencoded;charset=utf-8"},
    )
    response.raise_for_status()
    data = response.json()
    if int(data.get("code") or 0) != 200:
        raise ValueError(f"{path} failed: {data}")
    if data.get("success") is False:
        raise ValueError(f"{path} failed: {data}")
    return data


async def fetch_admission_site_payload(client: httpx.AsyncClient) -> dict[str, Any]:
    content_categories_payload = await post_form_api(
        client,
        "/api/front/commcfg/getXwfl",
        {"reqchannel": "site"},
    )
    content_categories = list_payload(content_categories_payload.get("data"))
    articles: list[dict[str, Any]] = []
    for category in content_categories:
        category_id = clean(category.get("id"))
        if not category_id:
            continue
        page_no = 1
        page_size = 50
        while True:
            list_payload_data = await post_form_api(
                client,
                "/api/front/commcfg/getXwListByFl",
                {
                    "reqchannel": "site",
                    "xwflid": category_id,
                    "pageNo": page_no,
                    "pageSize": page_size,
                },
            )
            data = dict_payload(list_payload_data.get("data"))
            article_list = list_payload(data.get("list"))
            for article in article_list:
                article_id = clean(article.get("id"))
                if not article_id:
                    continue
                detail_payload = await post_form_api(
                    client,
                    "/api/front/commcfg/getXwnrById",
                    {"reqchannel": "site", "id": article_id},
                )
                detail = dict_payload(detail_payload.get("data"))
                merged = {**article, **detail}
                merged["xwfl"] = article.get("xwfl") or detail.get("xwfl") or category
                articles.append(merged)
            count = as_int(data.get("count")) or len(article_list)
            if page_no * page_size >= count or not article_list:
                break
            page_no += 1

    academic_units_payload = await post_form_api(
        client,
        "/api/front/commcfg/getXuebuList",
        {"reqchannel": "site"},
    )
    media_categories_payload = await post_form_api(
        client,
        "/api/front/commcfg/getYxfl",
        {"reqchannel": "site"},
    )
    media_categories = list_payload(media_categories_payload.get("data"))
    media_items: list[dict[str, Any]] = []
    for category in media_categories:
        category_id = clean(category.get("id"))
        if not category_id:
            continue
        page_no = 1
        page_size = 200
        while True:
            media_payload = await post_form_api(
                client,
                "/api/front/commcfg/getYxnrByFl",
                {
                    "reqchannel": "site",
                    "yxflid": category_id,
                    "pageNo": page_no,
                    "pageSize": page_size,
                },
            )
            data = dict_payload(media_payload.get("data"))
            item_list = list_payload(data.get("list"))
            for item in item_list:
                media_items.append({**item, "category": category})
            count = as_int(data.get("count")) or len(item_list)
            if page_no * page_size >= count or not item_list:
                break
            page_no += 1

    return {
        "content_categories": content_categories,
        "content_articles": articles,
        "academic_units": list_payload(academic_units_payload.get("data")),
        "media_categories": media_categories,
        "media_items": media_items,
    }


def iter_queries(
    type_map: dict[str, Any],
    *,
    data_type: str,
    province: str,
    year: str,
    subject_type: str,
) -> Iterable[dict[str, str]]:
    for key in sorted(type_map):
        parts = key.split("_")
        if len(parts) != 4:
            continue
        sf, nf, klmc, xqmc = parts
        if klmc == "全部":
            continue
        if province and sf != province:
            continue
        if year and nf != year:
            continue
        if subject_type and klmc != subject_type:
            continue
        yield {
            "type": data_type,
            "sf": sf,
            "nf": nf,
            "zslb": "全部",
            "klmc": klmc,
            "xqmc": "" if xqmc in {"全部", "本校区", "校本部"} else xqmc,
        }


def normalize_plan_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        normalized = {
            "space_id": SPACE_ID,
            "year": as_int(row.get("nf")),
            "province": clean(row.get("sf")),
            "subject_type": clean(row.get("klmc")),
            "batch": clean(row.get("zslb") or row.get("pcmc")),
            "major_name": clean(row.get("zymc")),
            "class_type": clean(row.get("zslb")),
            "plan_count": as_int(row.get("jhrs")),
            "tuition": clean(row.get("zyxf")),
            "schooling_years": clean(row.get("xzmc")),
            "remarks": plan_remarks(row),
            "source_url": SOURCE_URL,
            "source_document": "武汉大学本科招生计划查询",
            "source_text": plan_row_markdown(row_to_plan(normalized_source=row)),
            "source_department": "武汉大学本科招生办公室",
            "published_at": clean(row.get("nf")),
            "raw_json": row,
        }
        if not normalized["year"] or not normalized["province"] or not normalized["subject_type"] or not normalized["major_name"]:
            continue
        key = (
            normalized["space_id"],
            normalized["year"],
            normalized["province"],
            normalized["subject_type"],
            normalized["batch"],
            normalized["major_name"],
            normalized["class_type"],
        )
        unique[key] = normalized
    return list(unique.values())


def normalize_score_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        normalized = {
            "space_id": SPACE_ID,
            "year": as_int(row.get("nf")),
            "province": clean(row.get("sf")),
            "subject_type": clean(row.get("klmc")),
            "batch": clean(row.get("zslb") or row.get("pcmc")),
            "major_name": clean(row.get("zymc")),
            "min_score": as_float(row.get("zdf")),
            "max_score": as_float(row.get("zgf")),
            "avg_score": as_float(row.get("pjf")),
            "min_rank": as_int(row.get("zdfwc")),
            "source_url": SOURCE_URL,
            "source_document": "武汉大学历年分数查询",
            "source_text": score_row_markdown(row_to_score(normalized_source=row)),
            "source_department": "武汉大学本科招生办公室",
            "published_at": clean(row.get("nf")),
            "raw_json": row,
        }
        if not normalized["year"] or not normalized["province"] or not normalized["subject_type"] or not normalized["major_name"]:
            continue
        key = (
            normalized["space_id"],
            normalized["year"],
            normalized["province"],
            normalized["subject_type"],
            normalized["batch"],
            normalized["major_name"],
        )
        unique[key] = normalized
    return list(unique.values())


def normalize_strong_foundation_rows(
    path: Path,
    *,
    year_filter: str,
    province_filter: str,
    subject_type_filter: str,
) -> list[dict[str, Any]]:
    if subject_type_filter or (year_filter and year_filter != "2025") or not path.exists():
        return []
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    headers = [clean_header(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    province_index = headers.index("省份")
    program_index = headers.index(clean_header("数学与应用数学\n（智能科学）\n强基计划"))
    rows: list[dict[str, Any]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        province = clean(values[province_index])
        if not province or province_filter and province != province_filter:
            continue
        raw_cell = clean(values[program_index])
        min_score, min_rank = parse_score_rank_cell(raw_cell)
        if min_score is None:
            continue
        source_text = strong_foundation_row_markdown(
            {
                "year": 2025,
                "province": province,
                "program_name": STRONG_FOUNDATION_PROGRAM,
                "min_score": min_score,
                "min_rank": min_rank,
            }
        )
        rows.append(
            {
                "space_id": SPACE_ID,
                "year": 2025,
                "province": province,
                "program_name": STRONG_FOUNDATION_PROGRAM,
                "subject_type": "",
                "min_score": min_score,
                "min_rank": min_rank,
                "source_url": SOURCE_URL,
                "source_document": path.name,
                "source_text": source_text,
                "source_department": "武汉大学本科招生办公室",
                "published_at": "2025",
                "raw_json": {
                    "source_file": str(path),
                    "province": province,
                    "program_name": STRONG_FOUNDATION_PROGRAM,
                    "score_rank_cell": raw_cell,
                },
            }
        )
    workbook.close()
    return rows


def normalize_content_categories(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        category_id = clean(row.get("id"))
        name = clean(row.get("name"))
        if not category_id or not name:
            continue
        normalized.append(
            {
                "category_id": category_id,
                "space_id": SPACE_ID,
                "name": name,
                "sort_order": as_int(row.get("sort")),
                "source_url": f"{SITE_URL}#/news/{category_id}",
                "source_document": "武汉大学本科招生网内容栏目",
                "source_department": "武汉大学本科招生办公室",
                "published_at": published_date(row.get("updateDate") or row.get("createDate")),
                "raw_json": row,
            }
        )
    return normalized


def normalize_content_articles(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique: dict[str, dict[str, Any]] = {}
    for row in rows:
        article_id = clean(row.get("id"))
        title = clean(row.get("title"))
        if not article_id or not title:
            continue
        category = dict_payload(row.get("xwfl"))
        category_id = clean(category.get("id"))
        category_name = clean(category.get("name"))
        source_url = article_url(category_id=category_id, article_id=article_id)
        markdown = html_to_markdown(clean(row.get("content")), base_url="https://zsdata.whu.edu.cn/")
        if not markdown:
            markdown = clean(row.get("description"))
        unique[article_id] = {
            "article_id": article_id,
            "space_id": SPACE_ID,
            "category_id": category_id or None,
            "category_name": category_name,
            "title": title,
            "description": clean(row.get("description")),
            "source_url": source_url,
            "logo_url": absolute_site_url(row.get("logo")),
            "content_type": clean(row.get("type")),
            "published_at": published_date(row.get("fbsj") or row.get("updateDate") or row.get("createDate")),
            "view_count": as_int(row.get("viewCount")),
            "source_document": title,
            "source_department": "武汉大学本科招生办公室",
            "source_text": article_source_text(title=title, category=category_name, markdown=markdown),
            "markdown": markdown,
            "raw_json": row,
        }
    return list(unique.values())


def normalize_academic_structure(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    units: list[dict[str, Any]] = []
    schools: list[dict[str, Any]] = []
    majors: dict[str, dict[str, Any]] = {}
    source_url = f"{SITE_URL}#/xyzy"
    for unit in rows:
        unit_id = clean(unit.get("id"))
        unit_name = clean(unit.get("name"))
        if not unit_id or not unit_name:
            continue
        units.append(
            {
                "unit_id": unit_id,
                "space_id": SPACE_ID,
                "name": unit_name,
                "sort_order": as_int(unit.get("sort")),
                "source_url": source_url,
                "source_document": "武汉大学本科招生网学院专业",
                "source_department": "武汉大学本科招生办公室",
                "published_at": published_date(unit.get("updateDate") or unit.get("createDate")),
                "raw_json": unit,
            }
        )
        for school in list_payload(unit.get("xyList")):
            school_id = clean(school.get("id"))
            school_name = clean(school.get("name"))
            if not school_id or not school_name:
                continue
            schools.append(
                {
                    "school_id": school_id,
                    "space_id": SPACE_ID,
                    "unit_id": unit_id,
                    "unit_name": unit_name,
                    "name": school_name,
                    "official_url": clean(school.get("url")),
                    "logo_url": absolute_site_url(school.get("logo")),
                    "sort_order": as_int(school.get("sort")),
                    "source_url": source_url,
                    "source_document": "武汉大学本科招生网学院专业",
                    "source_department": "武汉大学本科招生办公室",
                    "published_at": published_date(school.get("updateDate") or school.get("createDate")),
                    "raw_json": school,
                }
            )
            for major in list_payload(school.get("zyList")):
                major_id = clean(major.get("id"))
                major_name = clean(major.get("name"))
                if not major_name:
                    continue
                source_text = f"{major_name} 属于 {school_name}，所在学部为 {unit_name}。"
                majors[major_name] = {
                    "space_id": SPACE_ID,
                    "name": major_name,
                    "school_name": school_name,
                    "degree": None,
                    "category": unit_name,
                    "source_url": source_url,
                    "source_document": "武汉大学本科招生网学院专业",
                    "source_text": source_text,
                    "source_department": "武汉大学本科招生办公室",
                    "published_at": published_date(major.get("updateDate") or major.get("createDate")),
                    "raw_json": {
                        **major,
                        "id": major_id,
                        "school_id": school_id,
                        "school_name": school_name,
                        "unit_id": unit_id,
                        "unit_name": unit_name,
                    },
                }
    return {"units": units, "schools": schools, "majors": list(majors.values())}


def normalize_media_categories(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        category_id = clean(row.get("id"))
        name = clean(row.get("name"))
        if not category_id or not name:
            continue
        normalized.append(
            {
                "category_id": category_id,
                "space_id": SPACE_ID,
                "name": name,
                "sort_order": as_int(row.get("sort")),
                "source_url": f"{SITE_URL}#/yxnr/{category_id}",
                "source_document": "武汉大学本科招生网影像栏目",
                "source_department": "武汉大学本科招生办公室",
                "published_at": None,
                "raw_json": row,
            }
        )
    return normalized


def normalize_media_items(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique: dict[str, dict[str, Any]] = {}
    for row in rows:
        item_id = clean(row.get("id"))
        title = clean(row.get("title") or row.get("name"))
        if not item_id or not title:
            continue
        category = dict_payload(row.get("category"))
        category_id = clean(category.get("id"))
        category_name = clean(category.get("name"))
        media_url = absolute_site_url(row.get("url") or row.get("path") or row.get("video"))
        source_url = f"{SITE_URL}#/yxnr?flid={category_id}&id={item_id}"
        description = html_to_markdown(clean(row.get("content")), base_url="https://zsdata.whu.edu.cn/")
        if not description:
            description = clean(row.get("description"))
        source_text = "，".join(
            part
            for part in (
                f"{category_name}：{title}" if category_name else title,
                description,
                f"媒体地址：{media_url}" if media_url else "",
            )
            if part
        )
        unique[item_id] = {
            "item_id": item_id,
            "space_id": SPACE_ID,
            "category_id": category_id,
            "category_name": category_name,
            "title": title,
            "item_type": clean(row.get("type")),
            "source_url": source_url,
            "media_url": media_url,
            "logo_url": absolute_site_url(row.get("logo")),
            "description": description,
            "published_at": published_date(row.get("fbsj") or row.get("updateDate") or row.get("createDate")),
            "source_document": title,
            "source_department": "武汉大学本科招生办公室",
            "source_text": source_text,
            "raw_json": row,
        }
    return list(unique.values())


def build_admission_entities(
    *,
    plans: list[dict[str, Any]],
    scores: list[dict[str, Any]],
    strong_foundation_scores: list[dict[str, Any]],
    site_majors: list[dict[str, Any]],
    seed_path: Path,
) -> dict[str, list[dict[str, Any]]]:
    entities: dict[str, dict[str, Any]] = {}
    aliases: dict[tuple[str, str], dict[str, Any]] = {}
    relations: dict[tuple[str, str, str], dict[str, Any]] = {}
    seed_keys: dict[str, str] = {}

    def add_entity(
        *,
        entity_type: str,
        canonical_name: str,
        space_id: str = SPACE_ID,
        description: str = "",
        source_collection: str = "",
        source_key: str = "",
        metadata: dict[str, Any] | None = None,
    ) -> str:
        entity_id = stable_entity_id(space_id, entity_type, canonical_name)
        entities[entity_id] = {
            "entity_id": entity_id,
            "space_id": space_id,
            "entity_type": entity_type,
            "canonical_name": canonical_name,
            "description": description,
            "source_collection": source_collection,
            "source_key": source_key,
            "metadata": metadata or {},
        }
        add_alias(entity_id, canonical_name, "official", 1.0)
        return entity_id

    def add_alias(entity_id: str, alias: str, alias_type: str, confidence: float) -> None:
        normalized = normalize_entity_text(alias)
        if not normalized:
            return
        entity = entities.get(entity_id)
        space_id = str(entity.get("space_id") or SPACE_ID) if entity else SPACE_ID
        aliases[(entity_id, normalized)] = {
            "entity_id": entity_id,
            "space_id": space_id,
            "alias": alias,
            "normalized_alias": normalized,
            "alias_type": alias_type,
            "confidence": confidence,
        }

    def add_relation(
        subject_id: str,
        predicate: str,
        object_id: str,
        confidence: float,
        *,
        space_id: str = SPACE_ID,
    ) -> None:
        relations[(subject_id, predicate, object_id)] = {
            "space_id": space_id,
            "subject_entity_id": subject_id,
            "predicate": predicate,
            "object_entity_id": object_id,
            "confidence": confidence,
        }

    seed_payload = load_entity_seed(seed_path)
    for seed in seed_payload.get("entities", []):
        if not isinstance(seed, dict):
            continue
        entity_id = add_entity(
            entity_type=clean(seed.get("entity_type")),
            canonical_name=clean(seed.get("canonical_name")),
            space_id=GLOBAL_ENTITY_SPACE_ID,
            description=clean(seed.get("description")),
            source_collection=clean(seed.get("source_collection")),
            source_key=clean(seed.get("source_key")),
            metadata=dict(seed.get("metadata") or {}) if isinstance(seed.get("metadata"), dict) else {},
        )
        if seed.get("key"):
            seed_keys[str(seed["key"])] = entity_id
        for alias_item in seed.get("aliases") or []:
            if not isinstance(alias_item, dict):
                continue
            add_alias(
                entity_id,
                clean(alias_item.get("alias")),
                clean(alias_item.get("alias_type")) or "alias",
                float(alias_item.get("confidence") or 1.0),
            )
    for relation in seed_payload.get("relations", []):
        if not isinstance(relation, dict):
            continue
        subject_id = seed_keys.get(str(relation.get("subject_key") or ""))
        object_id = seed_keys.get(str(relation.get("object_key") or ""))
        predicate = clean(relation.get("predicate"))
        if subject_id and object_id and predicate:
            add_relation(
                subject_id,
                predicate,
                object_id,
                float(relation.get("confidence") or 1.0),
                space_id=GLOBAL_ENTITY_SPACE_ID,
            )

    major_rows: list[dict[str, Any]] = [
        *({"major_name": row.get("major_name")} for row in plans + scores),
        *({"major_name": row.get("name"), "school_name": row.get("school_name")} for row in site_majors),
    ]
    for row in major_rows:
        major_name = clean(row.get("major_name"))
        if not major_name:
            continue
        major_id = add_entity(
            entity_type="major",
            canonical_name=major_name,
            source_collection="admissions",
            source_key=major_name,
            metadata={
                "fact_tables": ["admission_plans", "admission_scores"],
                "fact_column": "major_name",
                **({"school_name": clean(row.get("school_name"))} if clean(row.get("school_name")) else {}),
            },
        )
        add_alias(major_id, major_name.replace("（", "(").replace("）", ")"), "display_variant", 0.96)
        for group in re_parentheses(major_name):
            add_alias(major_id, group, "short_name", 0.72)
    for province in sorted({clean(row.get("province")) for row in plans + scores + strong_foundation_scores}):
        if not province:
            continue
        province_id = add_entity(
            entity_type="province",
            canonical_name=province,
            source_collection="admissions",
            source_key=province,
            metadata={"fact_column": "province"},
        )
        add_alias(province_id, province, "official", 1.0)

    if strong_foundation_scores:
        program_id = add_entity(
            entity_type="program",
            canonical_name=STRONG_FOUNDATION_PROGRAM,
            description="武汉大学数学与应用数学（智能科学）强基计划录取分数",
            source_collection="admission_strong_foundation_scores",
            source_key=STRONG_FOUNDATION_PROGRAM,
            metadata={
                "fact_table": "admission_strong_foundation_scores",
                "fact_column": "program_name",
                "metric": "min_score",
            },
        )
        for alias, confidence in (
            ("数学与应用数学智能科学强基计划", 0.98),
            ("智能科学强基计划", 0.95),
            ("智能科学强基", 0.92),
            ("数学智能科学强基", 0.9),
        ):
            add_alias(program_id, alias, "short_name", confidence)
        if seed_keys.get("strong_foundation"):
            add_relation(program_id, "is_a", seed_keys["strong_foundation"], 1.0)
        if seed_keys.get("sai"):
            add_relation(program_id, "related_to", seed_keys["sai"], 0.85)

    return {
        "entities": list(entities.values()),
        "aliases": list(aliases.values()),
        "relations": list(relations.values()),
    }


def build_search_items(
    *,
    entity_payload: dict[str, list[dict[str, Any]]],
    plans: list[dict[str, Any]],
    scores: list[dict[str, Any]],
    strong_foundation_scores: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    aliases_by_entity: dict[str, list[dict[str, Any]]] = {}
    for alias in entity_payload["aliases"]:
        aliases_by_entity.setdefault(str(alias["entity_id"]), []).append(alias)
    for entity in entity_payload["entities"]:
        entity_aliases = sorted(
            aliases_by_entity.get(str(entity["entity_id"]), []),
            key=lambda item: float(item.get("confidence") or 0.0),
            reverse=True,
        )
        alias_text = "，".join(str(alias["alias"]) for alias in entity_aliases[:20])
        content = "\n".join(
            part
            for part in (
                str(entity["canonical_name"]),
                f"类型：{entity['entity_type']}",
                f"别名：{alias_text}" if alias_text else "",
                f"描述：{entity.get('description')}" if entity.get("description") else "",
            )
            if part
        )
        items.append(
            {
                "item_id": stable_search_item_id(str(entity["space_id"]), "entity", str(entity["entity_id"])),
                "space_id": entity["space_id"],
                "item_type": "entity",
                "entity_id": entity["entity_id"],
                "title": str(entity["canonical_name"]),
                "content_text": content,
                "metadata": {
                    "entity_type": entity["entity_type"],
                    "canonical_name": entity["canonical_name"],
                    "description": entity.get("description") or "",
                    "entity_metadata": entity.get("metadata") or {},
                    "aliases": [alias["alias"] for alias in entity_aliases],
                },
            }
        )
    for row in plans:
        key = "|".join(
            str(row.get(field) or "")
            for field in ("year", "province", "subject_type", "batch", "major_name", "class_type")
        )
        items.append(
            {
                "item_id": stable_search_item_id(SPACE_ID, "fact", f"admission_plans:{key}"),
                "space_id": SPACE_ID,
                "item_type": "fact",
                "fact_table": "admission_plans",
                "fact_key": key,
                "title": "武汉大学本科招生计划",
                "content_text": plan_row_markdown(row),
                "metadata": row,
            }
        )
    for row in scores:
        key = "|".join(
            str(row.get(field) or "")
            for field in ("year", "province", "subject_type", "batch", "major_name")
        )
        items.append(
            {
                "item_id": stable_search_item_id(SPACE_ID, "fact", f"admission_scores:{key}"),
                "space_id": SPACE_ID,
                "item_type": "fact",
                "fact_table": "admission_scores",
                "fact_key": key,
                "title": "武汉大学历年录取分数",
                "content_text": score_row_markdown(row),
                "metadata": row,
            }
        )
    for row in strong_foundation_scores:
        key = "|".join(str(row.get(field) or "") for field in ("year", "province", "program_name"))
        items.append(
            {
                "item_id": stable_search_item_id(SPACE_ID, "fact", f"admission_strong_foundation_scores:{key}"),
                "space_id": SPACE_ID,
                "item_type": "fact",
                "fact_table": "admission_strong_foundation_scores",
                "fact_key": key,
                "title": "武汉大学强基计划录取分数",
                "content_text": strong_foundation_row_markdown(row),
                "metadata": row,
            }
        )
    return items


def load_entity_seed(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"entities": [], "relations": []}
    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, dict) else {"entities": [], "relations": []}


def row_to_plan(*, normalized_source: dict[str, Any]) -> dict[str, Any]:
    return {
        "year": normalized_source.get("nf"),
        "province": normalized_source.get("sf"),
        "subject_type": normalized_source.get("klmc"),
        "batch": normalized_source.get("zslb") or normalized_source.get("pcmc"),
        "major_name": normalized_source.get("zymc"),
        "plan_count": normalized_source.get("jhrs"),
        "tuition": normalized_source.get("zyxf"),
        "schooling_years": normalized_source.get("xzmc"),
        "remarks": plan_remarks(normalized_source),
    }


def row_to_score(*, normalized_source: dict[str, Any]) -> dict[str, Any]:
    return {
        "year": normalized_source.get("nf"),
        "province": normalized_source.get("sf"),
        "subject_type": normalized_source.get("klmc"),
        "batch": normalized_source.get("zslb") or normalized_source.get("pcmc"),
        "major_name": normalized_source.get("zymc"),
        "min_score": normalized_source.get("zdf"),
        "max_score": normalized_source.get("zgf"),
        "avg_score": normalized_source.get("pjf"),
        "min_rank": normalized_source.get("zdfwc"),
    }


def write_dataset_artifact(
    artifact_store: MarkdownArtifactStore,
    *,
    title: str,
    url: str,
    rows: list[dict[str, Any]],
    raw_payload: list[dict[str, Any]],
    row_markdown,
):
    body = dataset_markdown(title=title, rows=rows, row_markdown=row_markdown)
    quality = MarkdownQualityChecker().check(body).to_dict()
    return artifact_store.write_document(
        site_id=SITE_ID,
        space_id=SPACE_ID,
        url=url,
        title=title,
        published_at=current_year(rows),
        markdown_body=body,
        raw_html=json.dumps(raw_payload, ensure_ascii=False, indent=2),
        quality=quality,
        depth=0,
        links=[{"url": SOURCE_URL, "text": "武汉大学本科招生数据查询", "is_asset": False}],
    )


def write_article_artifacts(
    artifact_store: MarkdownArtifactStore,
    rows: list[dict[str, Any]],
):
    artifacts = []
    checker = MarkdownQualityChecker()
    for row in rows:
        markdown = clean(row.get("markdown")) or clean(row.get("source_text"))
        if not markdown:
            continue
        title = clean(row.get("title")) or "武汉大学本科招生网文章"
        body = "\n\n".join(
            part
            for part in (
                f"# {title}",
                f"栏目：{row.get('category_name')}" if row.get("category_name") else "",
                markdown,
            )
            if part
        )
        quality = checker.check(body).to_dict()
        artifact = artifact_store.write_document(
            site_id=SITE_ID,
            space_id=SPACE_ID,
            url=str(row["source_url"]),
            title=title,
            published_at=row.get("published_at"),
            markdown_body=body,
            raw_html=json.dumps(row.get("raw_json") or {}, ensure_ascii=False, indent=2),
            quality=quality,
            depth=1,
            links=article_links(row),
        )
        artifacts.append(artifact)
    return artifacts


def write_media_catalog_artifact(
    artifact_store: MarkdownArtifactStore,
    rows: list[dict[str, Any]],
):
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(clean(row.get("category_name")) or "未分类", []).append(row)
    lines = ["# 武汉大学本科招生网影像与专题内容", "", f"数据来源：[{SITE_URL}]({SITE_URL})", ""]
    for category, group in sorted(grouped.items()):
        lines.extend([f"## {category}", ""])
        for row in sorted(group, key=lambda item: clean(item.get("title"))):
            line = f"- {row.get('title')}"
            if row.get("description"):
                line += f"：{row.get('description')}"
            if row.get("media_url"):
                line += f"。媒体地址：{row.get('media_url')}"
            lines.append(line)
        lines.append("")
    body = "\n".join(lines).strip()
    return artifact_store.write_document(
        site_id=SITE_ID,
        space_id=SPACE_ID,
        url=f"{SITE_URL}#/yxnr",
        title="武汉大学本科招生网影像与专题内容",
        published_at=None,
        markdown_body=body,
        raw_html=json.dumps(rows, ensure_ascii=False, indent=2),
        quality=MarkdownQualityChecker().check(body).to_dict(),
        depth=0,
        links=[{"url": str(row.get("source_url") or ""), "text": str(row.get("title") or ""), "is_asset": False} for row in rows],
    )


def write_academic_structure_artifact(
    artifact_store: MarkdownArtifactStore,
    payload: dict[str, list[dict[str, Any]]],
):
    schools_by_unit: dict[str, list[dict[str, Any]]] = {}
    majors_by_school: dict[str, list[dict[str, Any]]] = {}
    for school in payload["schools"]:
        schools_by_unit.setdefault(clean(school.get("unit_name")), []).append(school)
    for major in payload["majors"]:
        majors_by_school.setdefault(clean(major.get("school_name")), []).append(major)
    lines = ["# 武汉大学本科招生网学院专业", "", f"数据来源：[{SITE_URL}#/xyzy]({SITE_URL}#/xyzy)", ""]
    for unit in sorted(payload["units"], key=lambda item: (as_int(item.get("sort_order")) or 9999, clean(item.get("name")))):
        unit_name = clean(unit.get("name"))
        lines.extend([f"## {unit_name}", ""])
        for school in sorted(schools_by_unit.get(unit_name, []), key=lambda item: clean(item.get("name"))):
            school_name = clean(school.get("name"))
            lines.append(f"### {school_name}")
            if school.get("official_url"):
                lines.append(f"学院官网：{school.get('official_url')}")
            for major in sorted(majors_by_school.get(school_name, []), key=lambda item: clean(item.get("name"))):
                lines.append(f"- {major.get('name')}")
            lines.append("")
    body = "\n".join(lines).strip()
    return artifact_store.write_document(
        site_id=SITE_ID,
        space_id=SPACE_ID,
        url=f"{SITE_URL}#/xyzy",
        title="武汉大学本科招生网学院专业",
        published_at=None,
        markdown_body=body,
        raw_html=json.dumps(payload, ensure_ascii=False, indent=2),
        quality=MarkdownQualityChecker().check(body).to_dict(),
        depth=0,
        links=[{"url": str(school.get("official_url") or ""), "text": str(school.get("name") or ""), "is_asset": False} for school in payload["schools"]],
    )


async def index_artifact(store: PostgresKnowledgeStore, artifact) -> None:
    await store.upsert_document(
        IndexedDocument(
            document_id=artifact.document_id,
            space_id=str(artifact.metadata["space_id"]),
            site_id=str(artifact.metadata["site_id"]),
            title=str(artifact.metadata["title"]),
            source_url=str(artifact.metadata["url"]),
            published_at=artifact.metadata.get("published_at"),
            content_hash=str(artifact.metadata["content_hash"]),
            markdown_path=str(artifact.markdown_path),
            raw_html_path=str(artifact.raw_html_path),
            quality=dict(artifact.metadata.get("quality") or {}),
            markdown=artifact.markdown,
        )
    )


def dataset_markdown(*, title: str, rows: list[dict[str, Any]], row_markdown) -> str:
    grouped: dict[tuple[int, str], list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault((int(row["year"]), str(row["province"])), []).append(row)
    lines = [f"# {title}", "", f"数据来源：[{SOURCE_URL}]({SOURCE_URL})", ""]
    for (year, province), group in sorted(grouped.items(), reverse=True):
        lines.extend([f"## {year} {province}", ""])
        for row in sorted(group, key=lambda item: (str(item.get("subject_type")), str(item.get("major_name")))):
            lines.append(f"- {row_markdown(row)}")
        lines.append("")
    return "\n".join(lines).strip()


def plan_row_markdown(row: dict[str, Any]) -> str:
    parts = [
        f"{row.get('year')}年",
        str(row.get("province")),
        str(row.get("subject_type")),
        str(row.get("batch")),
        f"{row.get('major_name')} 招生计划 {row.get('plan_count')} 人",
    ]
    if row.get("schooling_years"):
        parts.append(f"学制 {row.get('schooling_years')}")
    if row.get("tuition"):
        parts.append(f"学费 {row.get('tuition')} 元/年")
    if row.get("remarks"):
        parts.append(str(row.get("remarks")))
    return "，".join(part for part in parts if part and part != "None")


def score_row_markdown(row: dict[str, Any]) -> str:
    parts = [
        f"{row.get('year')}年",
        str(row.get("province")),
        str(row.get("subject_type")),
        str(row.get("batch")),
        f"{row.get('major_name')} 最低分 {row.get('min_score')}",
        f"最高分 {row.get('max_score')}",
        f"平均分 {row.get('avg_score')}",
    ]
    if row.get("min_rank"):
        parts.append(f"最低位次 {row.get('min_rank')}")
    return "，".join(part for part in parts if part and part != "None")


def strong_foundation_row_markdown(row: dict[str, Any]) -> str:
    parts = [
        f"{row.get('year')}年",
        str(row.get("province")),
        str(row.get("program_name")),
        f"最低分 {row.get('min_score')}",
    ]
    if row.get("min_rank"):
        parts.append(f"最低位次 {row.get('min_rank')}")
    return "，".join(part for part in parts if part and part != "None")


def plan_remarks(row: dict[str, Any]) -> str:
    remarks = []
    if clean(row.get("zybz")):
        remarks.append(f"包含专业：{clean(row.get('zybz'))}")
    if clean(row.get("xkyq")):
        remarks.append(f"选考要求：{clean(row.get('xkyq'))}")
    if clean(row.get("zygroup")):
        remarks.append(f"专业组：{clean(row.get('zygroup'))}")
    return "；".join(remarks)


def html_to_markdown(html: str, *, base_url: str) -> str:
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    lines: list[str] = []
    for element in soup.find_all(["h1", "h2", "h3", "h4", "p", "li", "tr", "img"]):
        name = element.name.lower()
        if name == "img":
            src = absolute_site_url(element.get("src"), base_url=base_url)
            alt = clean(element.get("alt")) or "图片"
            if src:
                lines.append(f"![{alt}]({src})")
            continue
        if name == "tr":
            cells = [normalize_spaces(cell.get_text(" ", strip=True)) for cell in element.find_all(["th", "td"])]
            if cells:
                lines.append(" | ".join(cell for cell in cells if cell))
            continue
        text = normalize_spaces(element.get_text(" ", strip=True))
        if not text:
            continue
        if name in {"h1", "h2"}:
            lines.append(f"## {text}")
        elif name in {"h3", "h4"}:
            lines.append(f"### {text}")
        elif name == "li":
            lines.append(f"- {text}")
        else:
            lines.append(text)
    if not lines:
        text = normalize_spaces(soup.get_text(" ", strip=True))
        return text
    return "\n\n".join(dedupe_keep_order(lines)).strip()


def article_source_text(*, title: str, category: str, markdown: str) -> str:
    parts = [title]
    if category:
        parts.append(f"栏目：{category}")
    if markdown:
        parts.append(markdown[:2000])
    return "\n".join(parts)


def article_links(row: dict[str, Any]) -> list[dict[str, Any]]:
    links = [{"url": SITE_URL, "text": "武汉大学本科招生网", "is_asset": False}]
    if row.get("logo_url"):
        links.append({"url": str(row["logo_url"]), "text": "封面图", "is_asset": True})
    raw = dict_payload(row.get("raw_json"))
    for key in ("url", "video"):
        url = absolute_site_url(raw.get(key))
        if url:
            links.append({"url": url, "text": key, "is_asset": True})
    return links


def article_url(*, category_id: str, article_id: str) -> str:
    if category_id:
        return f"{SITE_URL}#/de/{category_id}/{article_id}"
    return f"{SITE_URL}#/de/{article_id}"


def absolute_site_url(value: Any, *, base_url: str = "https://zsdata.whu.edu.cn/") -> str:
    text = clean(value)
    if not text:
        return ""
    return urljoin(base_url, text)


def published_date(value: Any) -> str | None:
    text = clean(value)
    match = re.search(r"(20\d{2})[-/年.](\d{1,2})[-/月.](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    year = re.search(r"(20\d{2})", text)
    return year.group(1) if year else None


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def dedupe_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def dict_payload(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def list_payload(value: Any) -> list[dict[str, Any]]:
    return [item for item in value if isinstance(item, dict)] if isinstance(value, list) else []


def current_year(rows: list[dict[str, Any]]) -> str | None:
    years = sorted({int(row["year"]) for row in rows if row.get("year")}, reverse=True)
    return str(years[0]) if years else None


def clean(value: Any) -> str:
    return str(value or "").strip()


def as_int(value: Any) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def as_float(value: Any) -> float | None:
    text = clean(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_score_rank_cell(value: Any) -> tuple[float | None, int | None]:
    text = clean(value)
    if not text:
        return None, None
    parts = text.replace("／", "/").split("/")
    score = as_float(parts[0])
    rank = as_int(parts[1]) if len(parts) > 1 else None
    return score, rank


def clean_header(value: Any) -> str:
    return "".join(str(value or "").split())


def re_parentheses(value: str) -> list[str]:
    groups = []
    for group in re.findall(r"[（(]([^（）()]+)[）)]", value):
        text = group.strip()
        normalized = normalize_entity_text(text)
        if len(normalized) < 3 or re.fullmatch(r"\d+年?", normalized):
            continue
        groups.append(text)
    return groups


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


if __name__ == "__main__":
    asyncio.run(main())
